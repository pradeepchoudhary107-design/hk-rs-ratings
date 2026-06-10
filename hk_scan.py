import yfinance as yf
import pandas as pd
import requests
import json
import io
import os
import time
import smtplib
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

EMAIL_SENDER   = os.environ.get("EMAIL_SENDER", "")
EMAIL_PASSWORD = os.environ.get("EMAIL_PASSWORD", "")
EMAIL_TO       = os.environ.get("EMAIL_TO", "")

def get_hk_symbols():
    print("Fetching Hong Kong stock list...")
    all_symbols = []

    # HK stocks on Yahoo Finance use .HK suffix
    # HKEX lists ~2500+ stocks — numeric codes 0001 to 9999
    # Common range: 0001-4999 (main board) + 6000-9999 (some segments)
    # We generate the common ones and let yfinance validate
    print("  Generating HKEX symbol list (0001-4999 + selected range)...")

    # Known active HK stocks — major ones hardcoded as seed
    known = [
        "0700","0005","0941","1299","0388","2318","0939","1398","3988","0883",
        "0011","0001","0002","0003","0004","0006","0012","0016","0017","0019",
        "0023","0027","0066","0083","0101","0151","0175","0267","0291","0293",
        "0330","0358","0386","0390","0392","0669","0688","0762","0823","0857",
        "0868","0881","0960","0981","1038","1044","1093","1109","1113","1177",
        "1211","1288","1336","1339","1378","1810","1876","1928","1997","2007",
        "2018","2020","2269","2313","2319","2328","2333","2382","2388","2518",
        "2628","2688","3690","3692","3968","6098","6160","6690","6862","9618",
        "9888","9961","9988","9999","0003","0008","0013","0014","0020","0025",
        "0031","0033","0036","0038","0041","0045","0050","0051","0052","0053",
        "0054","0055","0056","0057","0058","0059","0060","0062","0063","0065",
        "0068","0069","0070","0071","0072","0073","0075","0076","0077","0078",
        "0079","0080","0081","0082","0084","0085","0086","0087","0088","0089",
        "0090","0091","0092","0093","0094","0095","0096","0097","0098","0099",
        "0100","0102","0103","0104","0105","0106","0107","0108","0110","0111",
        "0112","0113","0114","0115","0116","0117","0118","0119","0120","0121",
        "0122","0123","0124","0125","0126","0127","0128","0129","0130","0131",
        "0132","0133","0135","0136","0137","0138","0139","0140","0141","0142",
        "0143","0144","0145","0146","0147","0148","0150","0152","0153","0154",
        "0155","0156","0157","0158","0159","0160","0161","0163","0164","0165",
        "0166","0168","0169","0171","0172","0173","0174","0176","0177","0178",
        "0179","0180","0182","0183","0184","0185","0186","0187","0188","0189",
        "0190","0191","0193","0194","0195","0196","0197","0198","0199","0200",
    ]

    # Format as Yahoo Finance symbols: XXXX.HK
    symbols = [f"{s}.HK" for s in known]
    print(f"  {len(symbols)} HK symbols to scan")
    return symbols


def calc_rs(sym, start, end):
    try:
        ticker = yf.Ticker(sym)
        hist   = ticker.history(start=start, end=end, auto_adjust=True)
        if hist.empty or len(hist) < 150:
            return None
        p = hist["Close"].squeeze()
        if len(p) < 63:
            return None

        c_now = float(p.iloc[-1])
        c_3m  = float(p.iloc[-63])
        c_6m  = float(p.iloc[-126]) if len(p) >= 126 else float(p.iloc[0])
        c_9m  = float(p.iloc[-189]) if len(p) >= 189 else float(p.iloc[0])
        c_12m = float(p.iloc[-252]) if len(p) >= 252 else float(p.iloc[0])

        q4 = (c_now - c_3m)  / c_3m
        q3 = (c_3m  - c_6m)  / c_6m
        q2 = (c_6m  - c_9m)  / c_9m
        q1 = (c_9m  - c_12m) / c_12m
        score = 0.40*q4 + 0.20*q3 + 0.20*q2 + 0.20*q1

        high_52w  = float(p.iloc[-252:].max()) if len(p) >= 252 else float(p.max())
        ret_1m    = round(((c_now - float(p.iloc[-21])) / float(p.iloc[-21]))*100, 1) if len(p) >= 21 else 0
        ret_3m    = round(((c_now - c_3m) / c_3m)*100, 1)
        ret_12m   = round(((c_now - c_12m) / c_12m)*100, 1)
        from_high = round(((c_now - high_52w) / high_52w)*100, 1)

        try:
            info     = ticker.info
            mktcap_hkd = round(info.get("marketCap", 0) / 1e8, 1) if info.get("marketCap") else None
            sector   = info.get("sector",   "N/A")
            industry = info.get("industry", "N/A")
            name     = info.get("longName", sym.replace(".HK",""))
        except Exception:
            mktcap_hkd = None
            sector     = "N/A"
            industry   = "N/A"
            name       = sym.replace(".HK","")

        return {
            "sym":      sym.replace(".HK", ""),
            "name":     name,
            "score":    score,
            "price":    round(c_now, 2),
            "mktcap":   mktcap_hkd,
            "sector":   sector,
            "industry": industry,
            "r1m":      ret_1m,
            "r3m":      ret_3m,
            "r12m":     ret_12m,
            "h52":      round(high_52w, 2),
            "fh":       from_high,
        }
    except Exception:
        return None


def build_excel(df, fname):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    cols = [
        "Rank","Code","Company Name","RS Rating","Strength",
        "Sector","Industry","Mkt Cap (HKD Cr)",
        "Price (HKD)","1M Ret%","3M Ret%","12M Ret%",
        "52W High","From 52W High%"
    ]
    display = df.rename(columns={
        "rank":     "Rank",
        "sym":      "Code",
        "name":     "Company Name",
        "rs":       "RS Rating",
        "strength": "Strength",
        "sector":   "Sector",
        "industry": "Industry",
        "mktcap":   "Mkt Cap (HKD Cr)",
        "price":    "Price (HKD)",
        "r1m":      "1M Ret%",
        "r3m":      "3M Ret%",
        "r12m":     "12M Ret%",
        "h52":      "52W High",
        "fh":       "From 52W High%",
    })[cols]

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        display.to_excel(writer, sheet_name="HK RS Ratings", index=False)
        ws = writer.sheets["HK RS Ratings"]

        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="111111")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center")

        for row_idx in range(2, len(display) + 2):
            rs_val = ws.cell(row=row_idx, column=4).value
            if   rs_val >= 90: bg, fg = "085041", "9FE1CB"
            elif rs_val >= 80: bg, fg = "27500A", "C0DD97"
            elif rs_val >= 60: bg, fg = "633806", "FAC775"
            else:              bg, fg = "791F1F", "F7C1C1"
            for col in range(1, len(cols) + 1):
                c = ws.cell(row=row_idx, column=col)
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(color=fg, size=10)
                c.alignment = Alignment(horizontal="center")

        widths = [6, 8, 28, 10, 13, 20, 25, 16, 12, 10, 10, 12, 12, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    print(f"  Excel saved: {fname}")


def send_email(excel_path, df, date_str):
    if not EMAIL_PASSWORD:
        print("  EMAIL_PASSWORD not set — skipping.")
        return

    exc_count = len(df[df["rs"] >= 90])
    str_count = len(df[(df["rs"] >= 80) & (df["rs"] < 90)])
    subject   = f"HK Market RS Rating — {date_str} | {exc_count} Exceptional stocks"

    top10_rows = ""
    for _, row in df.head(10).iterrows():
        if row["rs"] >= 90:   bg, fg = "#085041", "#9FE1CB"
        elif row["rs"] >= 80: bg, fg = "#27500A", "#C0DD97"
        else:                 bg, fg = "#633806", "#FAC775"
        star     = "★" if row["rs"] >= 90 else "◆"
        c12      = "#1D9E75" if row["r12m"] >= 0 else "#E24B4A"
        mktcap_s = f"HK${row['mktcap']}Cr" if row["mktcap"] else "N/A"
        top10_rows += (
            "<tr>"
            f'<td style="padding:8px 10px;border-bottom:1px solid #222">{int(row["rank"])}</td>'
            f'<td style="padding:8px 10px;border-bottom:1px solid #222;font-weight:600">{row["sym"]}</td>'
            f'<td style="padding:8px 10px;border-bottom:1px solid #222;font-size:11px;color:#aaa">{row["name"][:25]}</td>'
            f'<td style="padding:8px 10px;border-bottom:1px solid #222">'
            f'<span style="background:{bg};color:{fg};padding:2px 8px;border-radius:5px;font-weight:700">'
            f'{row["rs"]} {star}</span></td>'
            f'<td style="padding:8px 10px;border-bottom:1px solid #222;color:#aaa;font-size:11px">{row["sector"]}</td>'
            f'<td style="padding:8px 10px;border-bottom:1px solid #222;color:{c12}">{row["r12m"]}%</td>'
            "</tr>"
        )

    body_html = f"""
    <div style="font-family:-apple-system,sans-serif;background:#0a0a0a;padding:2rem;max-width:680px;margin:0 auto">
      <div style="background:#111;border-radius:12px;padding:1.5rem;margin-bottom:1rem">
        <h1 style="color:#fff;font-size:1.2rem;font-weight:500;margin:0 0 4px">Hong Kong Market RS Rating</h1>
        <p style="color:#555;font-size:13px;margin:0">{date_str} · HKEX Listed Stocks</p>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:1rem">
        <div style="background:#111;border-radius:8px;padding:1rem;text-align:center">
          <div style="color:#555;font-size:11px;margin-bottom:4px">TOTAL SCANNED</div>
          <div style="color:#fff;font-size:24px;font-weight:500">{len(df)}</div>
        </div>
        <div style="background:#0a2218;border:1px solid #085041;border-radius:8px;padding:1rem;text-align:center">
          <div style="color:#1D9E75;font-size:11px;margin-bottom:4px">EXCEPTIONAL 90+</div>
          <div style="color:#9FE1CB;font-size:24px;font-weight:500">{exc_count}</div>
        </div>
        <div style="background:#0e1f07;border:1px solid #27500A;border-radius:8px;padding:1rem;text-align:center">
          <div style="color:#639922;font-size:11px;margin-bottom:4px">STRONG 80+</div>
          <div style="color:#C0DD97;font-size:24px;font-weight:500">{str_count}</div>
        </div>
      </div>
      <div style="background:#111;border-radius:12px;padding:1.5rem;margin-bottom:1rem">
        <h2 style="color:#aaa;font-size:13px;font-weight:500;margin:0 0 1rem;text-transform:uppercase">Top 10 HK Stocks</h2>
        <table style="width:100%;border-collapse:collapse;font-size:13px">
          <tr style="color:#555">
            <th style="text-align:left;padding:6px 10px;font-weight:500">#</th>
            <th style="text-align:left;padding:6px 10px;font-weight:500">Code</th>
            <th style="text-align:left;padding:6px 10px;font-weight:500">Company</th>
            <th style="text-align:left;padding:6px 10px;font-weight:500">RS</th>
            <th style="text-align:left;padding:6px 10px;font-weight:500">Sector</th>
            <th style="text-align:left;padding:6px 10px;font-weight:500">12M Ret</th>
          </tr>
          {top10_rows}
        </table>
      </div>
      <p style="color:#444;font-size:12px;text-align:center;margin:0">
        Full Excel attached · Code, Company Name, Sector, Industry, Mkt Cap · Auto every Saturday 10 AM IST
      </p>
    </div>"""

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = EMAIL_SENDER
    msg["To"]      = EMAIL_TO
    msg.attach(MIMEText(body_html, "html"))

    with open(excel_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(excel_path)}"')
    msg.attach(part)

    try:
        recipients = [e.strip() for e in EMAIL_TO.split(",")]
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, recipients, msg.as_string())
        print(f"  Email sent to: {recipients}")
    except Exception as e:
        print(f"  Email failed: {e}")


def main():
    t0       = time.time()
    date_str = datetime.now().strftime("%d %b %Y")
    print(f"HK RS Scan — {date_str}")

    symbols = get_hk_symbols()
    end     = datetime.today()
    start   = end - timedelta(days=420)
    results = []
    failed  = []

    for i, sym in enumerate(symbols):
        r = calc_rs(sym, start, end)
        if r:
            results.append(r)
        else:
            failed.append(sym)
        if (i + 1) % 50 == 0:
            elapsed = time.time() - t0
            rem     = (elapsed / (i + 1)) * (len(symbols) - i - 1) / 60
            print(f"  [{i+1}/{len(symbols)}] {len(results)} ok | {len(failed)} skipped | ~{rem:.0f} min left")
        time.sleep(0.3)

    if not results:
        print("No results.")
        return

    df       = pd.DataFrame(results)
    df["rs"] = df["score"].rank(pct=True).apply(
        lambda p: max(1, min(99, round(1 + p * 98)))
    ).astype(int)
    df       = df.sort_values("rs", ascending=False).drop(columns=["score"])
    df["rank"] = range(1, len(df) + 1)
    df["strength"] = df["rs"].apply(
        lambda r: "Exceptional" if r >= 90 else (
            "Strong" if r >= 80 else (
                "Average" if r >= 60 else "Weak"
            )
        )
    )

    records = df.to_dict(orient="records")
    payload = {
        "updated":     datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "updated_ist": datetime.now().strftime("%d %b %Y %I:%M %p IST"),
        "total":       len(records),
        "stocks":      records,
    }
    os.makedirs("docs", exist_ok=True)
    with open("docs/data.json", "w") as f:
        json.dump(payload, f, separators=(",", ":"))
    print(f"  JSON saved: {len(records)} stocks")

    ts = datetime.now().strftime("%Y%m%d")
    xl = f"HK_RS_{ts}.xlsx"
    build_excel(df, xl)
    send_email(xl, df, date_str)

    print(f"\nDone! {len(results)} stocks | {len(failed)} skipped | {round((time.time()-t0)/60,1)} min")

if __name__ == "__main__":
    main()
